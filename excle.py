# import openpyxl module
import openpyxl

wb = openpyxl.Workbook()

sheet = wb.active
"""
c1 = sheet.cell(row=1, column=1)

c1.value = "ANKIT"

c2 = sheet.cell(row=1, column=2)
c2.value = "RAI"

# Once have a Worksheet object, one can
# access a cell object by its name also.
# A2 means column = 1 & row = 2.
c3 = sheet['A2']
c3.value = "RAHUL"

# B2 means column = 2 & row = 2.
c4 = sheet['B2']
c4.value = "RAI"

"""
# Anytime you modify the Workbook object
# or its sheets and cells, the spreadsheet
# file will not be saved until you call
# the save() workbook method.

for i in range(1,5):
    if i <3:
        for j in range(1,5):
            c1 = sheet.cell(row=i, column=j)
            c1.value = "hello"+str(i)+str(j)


wb.save("C:\\Users\\ICB_AP\\PycharmProjects\\banglaText\\data\\main-data\\demo.xlsx")































"""

# Writing to an excel
# sheet using Python
import xlwt
from xlwt import Workbook

# Workbook is created
wb = Workbook()

# add_sheet is used to create sheet.
sheet1 = wb.add_sheet('Sheet 1')

sheet1.write(1, 0, 'ISBT DEHRADUN')
sheet1.write(2, 0, 'SHASTRADHARA')
sheet1.write(3, 0, 'CLEMEN TOWN')
sheet1.write(4, 0, 'RAJPUR ROAD')
sheet1.write(5, 0, 'CLOCK TOWER')


sheet1.write(0, 1, 'ISBT DEHRADUN')
sheet1.write(0, 2, 'SHASTRADHARA')
sheet1.write(0, 3, 'CLEMEN TOWN')
sheet1.write(0, 4, 'RAJPUR ROAD')
sheet1.write(0, 5, 'CLOCK TOWER')

wb.save('xlwt example.xls')





























"""

# Program to extract a particular row value
"""
import xlrd
import functionPython

dataParameter = "data/CCD-CCS/CCID.xlsx"
listOfcCDcCSWord = functionPython.LoadExcle(dataParameter)
dataParameter = "data/Adjective-Adverb/exel/jj-jq.xlsx"
listOfJJJQCSWord = functionPython.LoadExcle(dataParameter)

dataParameter = "data/negative-word/neg.txt"
listOfNegWord = functionPython.LoadData(dataParameter)


count = 0
flag = ""

for i in range(0, len(listOfNegWord)):
    if listOfNegWord[i] == "নয়":
        flag = "True"
        count = count+1
        #print("-Counted Negative - Word: " + str(count))
        break
    else:
        flag = "N/F"

print(flag)
"""
#score = functionPython.JJJQData(listOfJJJQCSWord, "অল্প")


#print(score)






"""

















"""
#loc = ("data/CCD-CCS/CCID.xlsx")
"""
loc = ("data/CCD-CCS/CCID.xlsx")
wb = xlrd.open_workbook(loc)
sheet = wb.sheet_by_index(0)

print(sheet.cell_value(0, 1))

listOfSentence = []

for i in range(sheet.nrows):
    data = sheet.cell_value(i, 0)
    listOfSentence.append(data)
    #for j in range(0, len(data)):

"""
"""
print(listOfSentence)

search = "পুনশ্চ"

for i in range(sheet.nrows):
    if search == sheet.cell_value(i, 0):
        datavalue = sheet.cell_value(i, 1)
        print(datavalue)
        break

"""


"""
#print(sheet.row_values(1))

"""