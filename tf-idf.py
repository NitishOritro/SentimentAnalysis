# Program to extract a particular row value
#
#NITISH RANJAN BHOWMIK
#
#Program Build a tf-IDF Model


import functionPython
from collections import Counter
import xlrd
import posTagger
import heapq
import xlwt
from xlwt import Workbook
import openpyxl
import numpy as np

from bnltk.stemmer import BanglaStemmer
from bnltk.tokenize import Tokenizers
t = Tokenizers()
fullStop = "।"

#Load Main Data

loc = ("data/main-data/Restaurant.xlsx")
wb = xlrd.open_workbook(loc)
sheet = wb.sheet_by_index(0)

sheet.cell_value(0, 0)


"""Load Dataset """

dataParameter = "data/Lexicon Dictionary Data/Resturant/correctPositive.txt"
listOfPositiveWord = functionPython.LoadData(dataParameter)
dataParameter = "data/Lexicon Dictionary Data/Resturant/correctNegative.txt"
listOfNegativeWord = functionPython.LoadData(dataParameter)

listOfTotalWord = listOfPositiveWord + listOfNegativeWord



sentence = ""
listOfSentence = []
listOfTotalSentence = []

for i in range(1,2059):
    data = sheet.cell_value(i, 1)
    for j in range(0, len(data)):
        lenData = len(data)
        if j == len(data)-1:
            lenData = len(data)
            sentence = sentence + data[j]
            listOfSentence.append(sentence)
            sentence = ""
            break
        else:
            sentence = sentence + data[j]
    #print(listOfSentence)
    listOfTotalSentence.append(listOfSentence)
    listOfSentence = []
print(listOfTotalSentence)


wordToCount = {}

checkWord = ""
for i in range(0, len(listOfTotalSentence)):
    extractToken = t.bn_word_tokenizer(listOfTotalSentence[i][0])
    #print(listOfTotalSentence[i])
    for word in extractToken:
        checkWord = functionPython.findWordFromList(listOfTotalWord, word)
        if checkWord == "True":
            checkWord = ""
            if word not in wordToCount.keys():
                wordToCount[word] = 1
            else:
                wordToCount[word] += 1

#print(wordToCount)

#print(wordToCount.items())
wordToList = []

for key, value in wordToCount.items():
    temp = [key,value]
    wordToList.append(temp)
print(wordToList)


wb = openpyxl.Workbook()
sheet = wb.active

c1 = sheet.cell(row=1, column=1)
c1.value = wordToList[0][0]

c2 = sheet.cell(row=1, column=2)
c2.value = wordToList[0][1]
for i in range(1, len(wordToList)):
    c1 = sheet.cell(row=i+1, column=1)
    c1.value = wordToList[i][0]

    c2 = sheet.cell(row=i + 1, column=2)
    c2.value = wordToList[i][1]

#wb.save("C:\\Users\\ICB_AP\\PycharmProjects\\banglaText\\data\\main-data\\dataWord.xlsx")
wb.save("C:\\PycharmProjects\\SentimentAnalysis\\data\\main-data\\dataWord.xlsx")


#idf matrix

word_idfs = {}
lengthDataset = len(listOfTotalSentence)

listOfWordWithTFDFvalue = [()]

#freq_words = heapq.nlargest(100,wordToCount,key=wordToCount.get(0))

#print(sorted(wordToCount.items(), key=lambda x: x[1], reverse=True))
wordToCount = (sorted(wordToCount.items(), key=lambda x: x[1], reverse=True))
k = Counter(wordToCount)

# Finding 3 highest values
wordToCountFre = k.most_common(10)

print(wordToCountFre)

for word in wordToCountFre:
    doc_count = 0
    for data in listOfTotalSentence:
        extractToken = t.bn_word_tokenizer(data[0])                        #extracttoken is a list
        #print(word[0]+" "+str(extractToken))
        if word[0][0] in extractToken:                                     #Word is tupple
            doc_count += 1
    word_idfs[word[0][0]] = np.log((lengthDataset / doc_count) + 1)
    #print(str(doc_count)+" "+str(word[0]))
print(word_idfs)

tf_matrix = {}
for word in wordToCountFre:
    doc_tf = []
    for data in listOfTotalSentence:
        frequency = 0
        extractToken = t.bn_word_tokenizer(data[0])
        for w in extractToken:
            if word[0][0] == w:
                frequency += 1
        tf_word = frequency / len(extractToken)
        doc_tf.append(tf_word)
    tf_matrix[word[0][0]] = doc_tf


#print(tf_matrix)

print(tf_matrix.keys())
print(tf_matrix)

#TF-IDF Calculation
tfIdf_matrix = []
for word in tf_matrix.keys():
    tfidf = []
    for tf in tf_matrix[word]:
        score = tf * word_idfs[word]
        tfidf.append(score)
    tfIdf_matrix.append(tfidf)

#print(tfIdf_matrix)


functionPython.SaveModelData(tfIdf_matrix, tf_matrix)


X = np.asarray(tfIdf_matrix)

X = np.transpose(X)

print(X)

import pandas as pd

## convert your array into a dataframe
df = pd.DataFrame (X)

## save to xlsx file

#filepath = 'C:\\Users\\ICB_AP\\PycharmProjects\\banglaText\\data\\main-data\\TransPosedataValue.xlsx'
filepath = 'C:\\PycharmProjects\\SentimentAnalysis\\data\\main-data\\TransPosedataValue.xlsx'

df.to_excel(filepath, index=False)


"""
for word in wordToList:
    doc_count = 0
    for i in range(0, len(listOfTotalSentence)):
        extractToken = t.bn_word_tokenizer(listOfTotalSentence[i][0])
        for j in range(0, len(extractToken)):
            if word[0] in str(extractToken[j]):
                #print(listOfTotalSentence[i][0])
                doc_count += 1
    #word_idfs[word] = np.log((lengthDataset/doc_count)+1)
#listOfWordWithTFDFvalue.append(tuple([word[0], np.log((lengthDataset/doc_count)+1)]))
    print(word[0]+" "+str(round(np.log((lengthDataset/doc_count)+1))))

#print(listOfWordWithTFDFvalue)
"""








