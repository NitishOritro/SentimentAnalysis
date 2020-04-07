import xlrd
import openpyxl

# Function definition is here

def LoadData(dataParameter):

    with open(dataParameter, encoding="utf8") as myfile:
        data = myfile.read()
    myfile.close()

    sentence = ""
    listOfWord = []
    fullStop = "।"
    newLine = "\n"

    res = None
    for i in range(0, len(data)):
        if data[i] != fullStop:
            if data[i] == newLine:
                res = i
                listOfWord.append(sentence)
                sentence = ""
            else:
                sentence = sentence + data[i]
        else:
            break

    return listOfWord

# Function definition is here
def LoadPositiveNegativeData(listOfPositiveWord, listOfNegativeWord, findWord):

    score = 0
    count = 0
    # find a word
    #find = findWord
    flag = ""
    for i in range(0, len(listOfPositiveWord)):
        if listOfPositiveWord[i] == findWord:
            flag = "true"
            score = 1
            #print("++++++ Word: "+findWord)
            break
        else:
            flag = "N/F"
    if flag == "N/F":
        for i in range(0, len(listOfNegativeWord)):
            if listOfNegativeWord[i] == findWord:
                flag = "true"
                score = -1
                #print("------- Word: "+ findWord)
                break
            else:
                flag = "N/F"

    if flag == "N/F":
        score = -999

    return score


def CountNegativeData(listOfNegWord, findWord):

    count = 0
    flag = ""

    for i in range(0, len(listOfNegWord)):
        if listOfNegWord[i] == findWord:
            flag = "True"
            count = count+1
            #print("-Counted Negative - Word: " + str(count))
            break
        else:
            flag = "N/F"

    return flag

# Function definition is here
def LoadExcle(dataParameter):
    wb = xlrd.open_workbook(dataParameter)
    sheet = wb.sheet_by_index(0)

    sheet.cell_value(0, 0)

    listOfWord = [()]
    dataValue = 0
    for i in range(sheet.nrows):
        data = sheet.cell_value(i, 0)
        dataValue = sheet.cell_value(i, 1)
        #append(tuple([3, 4]))
        listOfWord.append(tuple([data, dataValue]))
        # for j in range(0, len(data)):

    #print(listOfWord)

    return listOfWord


def cCDcCSData(listOfcCDcCSWord, findWord):
    flag = ""
    wb = xlrd.open_workbook("data/CCD-CCS/CCID.xlsx")
    sheet = wb.sheet_by_index(0)
    value = 0
    index = 0
    for i in range(0, len(listOfcCDcCSWord)):
        if i !=0 and listOfcCDcCSWord[i][0] == findWord:
            flag = "True"

            value = listOfcCDcCSWord[i][1]
            #print("CCD-CCS word:" + findWord+ " and Value is: "+ str(value))
            break
        else:
            flag = "N/F"
            value = 1

    return value

def JJJQData(listOfcCDcCSWord, findWord):
    flag = ""

    value = 0
    index = 0
    for i in range(0, len(listOfcCDcCSWord)):
        if i !=0 and listOfcCDcCSWord[i][0] == findWord:
            #index = i
            flag = "True"
            value = listOfcCDcCSWord[i][1]
            #print("JJ-JQ word:"+findWord +" and Value is: "+ str(value))

            break
        else:
            flag = "N/F"
            value = 1

    return value



def SaveData(listOfSentenceScore):

    wb = openpyxl.Workbook()
    sheet = wb.active

    c1 = sheet.cell(row=1, column=1)
    c1.value = listOfSentenceScore[0]
    for i in range(1, len(listOfSentenceScore)):
        c1 = sheet.cell(row=i+1, column=1)
        c1.value = listOfSentenceScore[i]

    #wb.save("C:\\Users\\ICB_AP\\PycharmProjects\\banglaText\\data\\main-data\\save.xlsx")
    wb.save("C:\\PycharmProjects\\SentimentAnalysis\\data\main-data\\save.xlsx")
    return "True"


def SaveModelData(IfIdf_matrix, tf_matrix):

    wb = openpyxl.Workbook()
    sheet = wb.active
    #c1 = sheet.cell(row=1, column=1)
    #c1.value = IfIdf_matrix[0][0]

    #c2 = sheet.cell(row=1, column=2)
    #c2.value = IfIdf_matrix[0][1]

    counter = 1
    for word in tf_matrix.keys():
        c2 = sheet.cell(row=counter, column=1)
        c2.value = word
        counter = counter+1


    for i in range(0, len(IfIdf_matrix)):
        for j in range(0, len(IfIdf_matrix[i])):
            #c2 = sheet.cell(row=i+1, column=1)
            #c2.value = word

            c1 = sheet.cell(row=i+1, column=j+2)
            c1.value = IfIdf_matrix[i][j]



        #c2 = sheet.cell(row=i + 1, column=2)
        #c2.value = IfIdf_matrix[i][1]

    #wb.save("C:\\Users\\ICB_AP\\PycharmProjects\\banglaText\\data\\main-data\\dataWordValue1.xlsx")
    wb.save("C:\\PycharmProjects\\SentimentAnalysis\\data\main-data\\dataWordValue1.xlsx")



def readFromExcle(loc):

    sentence = ""
    listOfSentence = []
    listOfTotalSentence = []

    wb = xlrd.open_workbook(loc)
    sheet = wb.sheet_by_index(0)

    sheet.cell_value(0, 0)

    for i in range(1, 20):
        data = sheet.cell_value(i, 1)
        for j in range(0, len(data)):
            lenData = len(data)
            if j == len(data) - 1:
                lenData = len(data)
                sentence = sentence + data[j]
                listOfSentence.append(sentence)
                sentence = ""
                break
            else:
                sentence = sentence + data[j]
        # print(listOfSentence)
        listOfTotalSentence.append(listOfSentence)
        listOfSentence = []
    return listOfTotalSentence
    #print(listOfTotalSentence)



####Find a Word in a List###

def findWordFromList(listOfTotalWord, find):

    for i in range(0, len(listOfTotalWord)):
        if listOfTotalWord[i] == find:
            #print("found case match")
            flag = "True"
            break
        else:
            flag = "False"
            #print("not found")

    return flag




