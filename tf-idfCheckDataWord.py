# Program to extract a particular row value
#
#NITISH RANJAN BHOWMIK
#
#Program Build a tf-IDF Model


import functionPython
from collections import Counter
import xlrd
import posTagger
import heapq
import xlwt
from xlwt import Workbook
import openpyxl
import numpy as np

from bnltk.stemmer import BanglaStemmer
from bnltk.tokenize import Tokenizers
t = Tokenizers()
fullStop = "।"

#Load Main Data

loc = ("data/main-data/Cricket.xlsx")
wb = xlrd.open_workbook(loc)
sheet = wb.sheet_by_index(0)

sheet.cell_value(0, 0)


"""Load Dataset """

dataParameter = "data/Lexicon Dictionary Data/Cricket/correctPositive.txt"
listOfPositiveWord = functionPython.LoadData(dataParameter)
dataParameter = "data/Lexicon Dictionary Data/Cricket/correctNegative.txt"
listOfNegativeWord = functionPython.LoadData(dataParameter)

listOfTotalWord = listOfPositiveWord + listOfNegativeWord


sentence = ""
listOfSentence = []
listOfTotalSentence = []

for i in range(1,2980):
    data = sheet.cell_value(i, 2)
    for j in range(0, len(data)):
        lenData = len(data)
        if j == len(data)-1:
            lenData = len(data)
            sentence = sentence + data[j]
            listOfSentence.append(sentence)
            sentence = ""
            break
        else:
            sentence = sentence + data[j]
    #print(listOfSentence)
    listOfTotalSentence.append(listOfSentence)
    listOfSentence = []
print(listOfTotalSentence)


wordToCount = {}

checkWord = ""
for i in range(0, len(listOfTotalSentence)):
    extractToken = t.bn_word_tokenizer(listOfTotalSentence[i][0])
    #print(listOfTotalSentence[i])
    for word in extractToken:
        checkWord = functionPython.findWordFromList(listOfTotalWord, word)
        if checkWord == "True":
            checkWord = ""
            if word not in wordToCount.keys():
                wordToCount[word] = 1
            else:
                wordToCount[word] += 1


print(wordToCount)

#print(wordToCount.items())
wordToList = []

for key, value in wordToCount.items():
    temp = [key,value]
    wordToList.append(temp)
print(wordToList)

wordToListCheck = sorted(wordToList, key=lambda x: x[1], reverse=True)


wb = openpyxl.Workbook()
sheet = wb.active

c1 = sheet.cell(row=1, column=1)
c1.value = wordToListCheck[0][0]

c2 = sheet.cell(row=1, column=2)
c2.value = wordToListCheck[0][1]
for i in range(1, len(wordToList)):
    c1 = sheet.cell(row=i+1, column=1)
    c1.value = wordToListCheck[i][0]

    c2 = sheet.cell(row=i + 1, column=2)
    c2.value = wordToListCheck[i][1]

#wb.save("C:\\Users\\ICB_AP\\PycharmProjects\\banglaText\\data\\main-data\\dataWord.xlsx")
wb.save("C:\\PycharmProjects\\SentimentAnalysis\\data\\main-data\\dataWord.xlsx")


#idf matrix

word_idfs = {}
lengthDataset = len(listOfTotalSentence)

listOfWordWithTFDFvalue = [()]

#freq_words = heapq.nlargest(100,wordToCount,key=wordToCount.get(0))

#print(sorted(wordToCount.items(), key=lambda x: x[1], reverse=True))
wordToCount = (sorted(wordToCount.items(), key=lambda x: x[1], reverse=True))
k = Counter(wordToCount)

# Finding 3 highest values
wordToCountFre = k.most_common(1500)

print(wordToCountFre)